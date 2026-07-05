#!/usr/bin/env python3
"""Expand Excel content: re-tier existing biases and add 40 new biases + 120 challenges."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "cognitive_bias_trainer.xlsx"

sys.path.insert(0, str(Path(__file__).resolve().parent))

from expand_content_data_tier2_biases import TIER2_BIASES
from expand_content_data_tier2_challenges_a import TIER2_CHALLENGES
from expand_content_data_tier2_challenges_b import TIER2_CHALLENGES_B
from expand_content_data_tier3_biases import TIER3_BIASES
from expand_content_data_tier3_challenges_a import TIER3_CHALLENGES_A
from expand_content_data_tier3_challenges_b import TIER3_CHALLENGES_B

CHALLENGE_END_ROW = 181  # rows 2..181 = 180 challenges


def bias_row(bias: dict) -> list:
    return [
        bias["bias_id"],
        bias["name"],
        bias["definition"],
        bias["category"],
        bias["description"],
        bias["everyday_tip"],
        bias["common_contexts"],
        bias["tier"],
    ]


def challenge_row(ch: dict) -> list:
    opts = ch["options"]
    return [
        ch["challenge_id"],
        ch["bias_id"],
        ch["bias"],
        ch["statement"],
        opts[0],
        opts[1],
        opts[2],
        opts[3],
        ch["correct"],
        ch["explanation"],
        ch["reflection"],
        ch["why"],
        ch["context"],
        ch["difficulty"],
        ch["objective"],
    ]


def stats_formula(bias_cell: str, kind: str) -> str:
    base = f"'Challenge Library'!$B$2:$B${CHALLENGE_END_ROW}"
    diff = f"'Challenge Library'!$N$2:$N${CHALLENGE_END_ROW}"
    if kind == "count":
        return f"=COUNTIF({base},{bias_cell})"
    return f'=COUNTIFS({base},{bias_cell},{diff},"{kind}")'


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE)

    bias_ws = wb["Bias Library"]
    challenge_ws = wb["Challenge Library"]
    stats_ws = wb["Bias Statistics"]

    # Re-tier all existing biases to Tier 1
    for row in range(2, bias_ws.max_row + 1):
        if bias_ws.cell(row, 1).value:
            bias_ws.cell(row, 8).value = 1

    # Append new biases
    new_biases = TIER2_BIASES + TIER3_BIASES
    for bias in new_biases:
        bias_ws.append(bias_row(bias))

    # Append new challenges
    all_new_challenges = (
        TIER2_CHALLENGES
        + TIER2_CHALLENGES_B
        + TIER3_CHALLENGES_A
        + TIER3_CHALLENGES_B
    )
    assert len(all_new_challenges) == 120, f"Expected 120 challenges, got {len(all_new_challenges)}"

    for ch in all_new_challenges:
        challenge_ws.append(challenge_row(ch))

    # Rebuild Bias Statistics sheet with updated formula ranges
    if stats_ws.max_row > 1:
        stats_ws.delete_rows(2, stats_ws.max_row - 1)

    for row in range(2, bias_ws.max_row + 1):
        bias_id = bias_ws.cell(row, 1).value
        if not bias_id:
            continue
        stats_row = stats_ws.max_row + 1
        stats_ws.cell(stats_row, 1).value = bias_id
        stats_ws.cell(stats_row, 2).value = bias_ws.cell(row, 2).value
        stats_ws.cell(stats_row, 3).value = bias_ws.cell(row, 4).value
        bias_cell = f"A{stats_row}"
        stats_ws.cell(stats_row, 4).value = stats_formula(bias_cell, "count")
        stats_ws.cell(stats_row, 5).value = stats_formula(bias_cell, "Recognition")
        stats_ws.cell(stats_row, 6).value = stats_formula(bias_cell, "Application")
        stats_ws.cell(stats_row, 7).value = stats_formula(bias_cell, "Ambiguous")

    wb.save(SOURCE)
    print(f"Updated {SOURCE.name}")
    print(f"  Biases: {bias_ws.max_row - 1}")
    print(f"  Challenges: {challenge_ws.max_row - 1}")


if __name__ == "__main__":
    main()
