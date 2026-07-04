#!/usr/bin/env python3
"""
Convert cognitive_bias_trainer.xlsx → JSON files in data/.

Run from repo root:
  python3 second-thought/scripts/convert-excel.py
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "cognitive_bias_trainer.xlsx"
OUT = ROOT / "data"


def row_dict(headers: tuple, row: tuple) -> dict:
    return {
        headers[i]: row[i]
        for i in range(len(headers))
        if headers[i] is not None
    }


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

    # --- Bias Library ---
    bias_rows = list(wb["Bias Library"].iter_rows(values_only=True))
    bias_headers = bias_rows[0]
    biases = []
    for row in bias_rows[1:]:
        if not row or not row[0]:
            continue
        src = row_dict(bias_headers, row)
        biases.append(
            {
                "biasId": src["bias_id"],
                "name": src["Bias"],
                "definition": src["Definition"],
                "category": src["Category"],
                "tier": int(src.get("Tier") or 1),
                "description": src["Description"],
                "everydayTip": src["Everyday Tip"],
                "commonContexts": [
                    c.strip() for c in str(src["Common Contexts"]).split(",")
                ],
            }
        )

    # --- Categories ---
    cat_rows = list(wb["Categories"].iter_rows(values_only=True))
    cat_headers = cat_rows[0]
    categories = []
    for row in cat_rows[1:]:
        if not row or not row[0]:
            continue
        src = row_dict(cat_headers, row)
        categories.append(
            {"categoryId": src["category_id"], "name": src["Category"]}
        )

    # --- Challenge Library ---
    ch_rows = list(wb["Challenge Library"].iter_rows(values_only=True))
    ch_headers = ch_rows[0]
    challenges = []
    for row in ch_rows[1:]:
        if not row or not row[0]:
            continue
        src = row_dict(ch_headers, row)
        challenges.append(
            {
                "challengeId": src["challenge_id"],
                "biasId": src["bias_id"],
                "biasName": src["Bias"],
                "statement": src["Statement"],
                "options": {
                    "A": src["Option A"],
                    "B": src["Option B"],
                    "C": src["Option C"],
                    "D": src["Option D"],
                },
                "correctAnswer": str(src["Correct Answer"]).strip().upper(),
                "explanation": src["Explanation"],
                "reflectionQuestion": src["Reflection Question"],
                "whyHumansDoThis": src["Why Humans Do This"],
                "realWorldContext": src["Real World Context"],
                "difficulty": src["Difficulty"],
                "learningObjective": src["Learning Objective"],
            }
        )

    wb.close()

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "biases.json").write_text(
        json.dumps(biases, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT / "categories.json").write_text(
        json.dumps(categories, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT / "challenges.json").write_text(
        json.dumps(challenges, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    print(f"Wrote {len(biases)} biases, {len(challenges)} challenges, {len(categories)} categories")


if __name__ == "__main__":
    main()
