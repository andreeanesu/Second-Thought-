#!/usr/bin/env python3
"""Apply lighter copy rewrites to Challenge Library rows."""

from __future__ import annotations

import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "cognitive_bias_trainer.xlsx"

UPDATES = {
    "C006": {
        "Statement": "In a job interview, they ask for my salary first — and that number ends up anchoring the whole negotiation.",
        "Explanation": "The first number spoken becomes the reference for everything after.",
        "Reflection Question": "Have you ever felt stuck near a number you said early on?",
        "Why Humans Do This": "Early numbers are hard to let go of — even when new info suggests we should move on.",
    },
    "C007": {
        "Statement": "After a few plane-crash stories in the news, flying feels terrifying — even though stats say driving is riskier.",
        "Explanation": "Vivid stories feel more common than they are.",
        "Reflection Question": "Any fear you have that might come from memorable stories, not stats?",
        "Why Humans Do This": "Easy-to-recall events feel more likely. Our brain uses memory as a shortcut.",
    },
    "C008": {
        "Statement": "I saw one viral post about a rare vaccine side effect — and it stuck more than the disease risks I ignored.",
        "Explanation": "One vivid story outweighed the broader data.",
        "Reflection Question": "On health risks, do you check the data or trust the story that sticks?",
        "Why Humans Do This": "Emotional stories lodge in memory. Dry statistics often don't.",
    },
    "C009": {
        "Statement": "After a friend's house was burgled, burglaries suddenly feel common here — even though crime has been falling.",
        "Explanation": "One close story outweighed the wider trend.",
        "Reflection Question": "Has one person's story ever changed how safe you feel?",
        "Why Humans Do This": "Stories from people we know feel real and relevant — more than numbers on a page.",
    },
    "C010": {
        "Statement": "A colleague was late to the meeting. Everyone assumed they're lazy — not knowing their car broke down.",
        "Explanation": "Character got blamed before circumstances were considered.",
        "Reflection Question": "Next time someone's late — what situation might explain it?",
        "Why Humans Do This": "We don't see others' full context, so we fill the gap with personality labels.",
    },
    "C011": {
        "Statement": "Someone cut me off in traffic. I thought 'what a jerk' — without wondering if they had an emergency.",
        "Explanation": "The action was blamed on character, not situation.",
        "Reflection Question": "Have you ever driven badly because of your own bad day?",
        "Why Humans Do This": "It's faster to judge a stranger's character than imagine their unseen story.",
    },
    "C012": {
        "Statement": "A friend cancelled last minute again. I decided they're unreliable — maybe it was just a run of bad luck.",
        "Explanation": "One label stuck before a real pattern was clear.",
        "Reflection Question": "How many times before you'd call something a pattern, not bad luck?",
        "Why Humans Do This": "One clear story feels easier than holding uncertainty over time.",
    },
    "C013": {
        "Statement": "I passed because I'm smart — but failed last time because the questions were unfair.",
        "Explanation": "Success went to me; failure went to outside factors.",
        "Reflection Question": "Your last win and loss — did you explain them the same way?",
        "Why Humans Do This": "Crediting ourselves for wins and blaming outside forces for losses protects how we feel about ourselves.",
    },
    "C014": {
        "Statement": "We won because of great strategy. We lost because of a bad referee.",
        "Explanation": "The win was ours; the loss was someone else's fault.",
        "Reflection Question": "A loss you blamed entirely on something outside your control?",
        "Why Humans Do This": "Keeping a positive self-image feels better than sitting with our own shortcomings.",
    },
    "C015": {
        "Statement": "A manager took credit when a project worked — and blamed the team when a similar one failed.",
        "Explanation": "Success was personal; failure was shifted away.",
        "Reflection Question": "When your project fails — how much do you own versus blame others?",
        "Why Humans Do This": "Protecting our sense of competence is easier than sharing responsibility for failure.",
    },
    "C016": {
        "Statement": "I keep holding a failing stock because I've already put so much into it — hoping it'll bounce back.",
        "Explanation": "Past money spent is driving the choice, not future prospects.",
        "Reflection Question": "If you were deciding fresh today, would you still hold it?",
        "Why Humans Do This": "Walking away feels like admitting a mistake. Staying avoids that pain — even when it costs more.",
    },
    "C017": {
        "Statement": "I've spent five years in this relationship. I can't leave now — even though I'm unhappy.",
        "Explanation": "Time already spent is treated as a reason to stay.",
        "Reflection Question": "Would you start this relationship today, knowing what you know?",
        "Why Humans Do This": "Leaving can feel like wasted years were wasted for nothing. That's painful.",
    },
    "C018": {
        "Statement": "We keep eating a mediocre meal we already paid for — even though we could stop.",
        "Explanation": "Money already spent shouldn't change whether we keep eating.",
        "Reflection Question": "Does having paid change how much you 'should' finish?",
        "Why Humans Do This": "We mix up 'already spent' with 'what to do next' — even when they're separate decisions.",
    },
    "C019": {
        "Statement": "Nine compliments and one criticism in my review — and I obsessed over the criticism all week.",
        "Explanation": "One negative note outweighed nine positive ones.",
        "Reflection Question": "Last feedback you got — praise or criticism stuck more?",
        "Why Humans Do This": "Bad news used to signal danger. Our brains still give it extra weight.",
    },
    "C020": {
        "Statement": "News runs more bad stories than good ones — because negative headlines get more clicks.",
        "Explanation": "Outlets and audiences both lean toward negative news.",
        "Reflection Question": "Do you click negative headlines more than positive ones?",
        "Why Humans Do This": "Threats and bad news pull our attention harder than good news.",
    },
    "C021": {
        "Statement": "One critical comment from my partner stuck more than months of compliments.",
        "Explanation": "A single negative remark dominated the emotional memory.",
        "Reflection Question": "One comment that overshadowed a lot of good moments?",
        "Why Humans Do This": "Negative moments are processed more deeply and remembered more vividly.",
    },
    "C022": {
        "Statement": "A polished, articulate candidate seemed competent at the technical work too — without any proof.",
        "Explanation": "A good impression in one area spilled into another.",
        "Reflection Question": "Assumed someone was good at their job because they seemed confident?",
        "Why Humans Do This": "One overall impression is easier than judging each skill separately.",
    },
    "C023": {
        "Statement": "A celebrity endorsed a health product — so fans assumed it works, because they like the celebrity.",
        "Explanation": "Feelings about the person transferred to the product.",
        "Reflection Question": "Would you trust the claim as much from a stranger?",
        "Why Humans Do This": "We pass trust from people we like to things they're linked with.",
    },
    "C024": {
        "Statement": "A charismatic colleague got promoted fast — and people assumed all their ideas were good, untested ones included.",
        "Explanation": "Likeability stretched to ideas without evidence.",
        "Reflection Question": "How do you separate liking someone from trusting their untested ideas?",
        "Why Humans Do This": "A strong first impression is hard to update — even when proof is missing.",
    },
    "C025": {
        "Statement": "A new hire was late on day one. Colleagues decided they're unreliable and probably not hardworking.",
        "Explanation": "One slip became a whole personality judgment.",
        "Reflection Question": "One early mistake that colored how you saw someone?",
        "Why Humans Do This": "One flaw gets simplified into 'this person is generally bad' instead of staying specific.",
    },
    "C026": {
        "Statement": "Someone mispronounced a word in a talk — and the audience doubted whether they know the topic at all.",
        "Explanation": "A small slip was stretched into a competence judgment.",
        "Reflection Question": "Judged someone's expertise over an unrelated small mistake?",
        "Why Humans Do This": "Our brain wants one consistent story — and one flaw builds a negative one fast.",
    },
    "C027": {
        "Statement": "A public figure got one fact wrong — and viewers dismissed everything else they said, even the solid parts.",
        "Explanation": "One error discredited the whole message.",
        "Reflection Question": "When someone errs once, do you dismiss everything they say?",
        "Why Humans Do This": "One clear mistake feels like reliable evidence — even when it isn't about everything else.",
    },
    "C028": {
        "Statement": "Success books interview winning entrepreneurs — not the many who tried the same thing and failed.",
        "Explanation": "We only hear from survivors, so strategies look more reliable than they are.",
        "Reflection Question": "Reading a success story — do you ask who tried the same and failed?",
        "Why Humans Do This": "Failures are quieter, less visible, and less fun to talk about.",
    },
    "C029": {
        "Statement": "'These five habits made CEOs successful' — with no check on failed CEOs who had the same habits.",
        "Explanation": "Without a comparison group, the link to success is shaky.",
        "Reflection Question": "Without asking who failed, can you trust why someone succeeded?",
        "Why Humans Do This": "Success stories are celebrated. Quiet failures disappear from view.",
    },
    "C030": {
        "Statement": "Everyone still in the course says it's easy — but the people who struggled already dropped out.",
        "Explanation": "Only survivors are in the sample you're hearing from.",
        "Reflection Question": "Whose voices might be missing from the feedback you hear?",
        "Why Humans Do This": "We hear from who's left — and forget who already walked away.",
    },
    "C031": {
        "Statement": "After a stock crashed, I said 'I knew that would happen' — though I'd bought more the week before.",
        "Explanation": "Memory of what we believed shifted to match the outcome.",
        "Reflection Question": "Claimed you 'knew' something — despite acting the opposite before?",
        "Why Humans Do This": "Once we know the ending, it feels inevitable — and our memory bends to match.",
    },
    "C032": {
        "Statement": "After the loss, I insisted 'I knew we'd lose' — though I'd predicted a win the day before.",
        "Explanation": "Recollection of the prediction was rewritten after the fact.",
        "Reflection Question": "Could you find what you actually predicted — versus what you claim now?",
        "Why Humans Do This": "We rebuild past beliefs to fit what we know now.",
    },
    "C033": {
        "Statement": "After a project failed, the team said 'the risks were obvious from the start' — though nobody flagged them in planning.",
        "Explanation": "Risks look obvious only after failure is known.",
        "Reflection Question": "Were the 'obvious' risks written down before the outcome?",
        "Why Humans Do This": "Knowing how it ended makes the path feel more predictable than it was.",
    },
    "C034": {
        "Statement": "I assume most people share my political view — because everyone in my circle does.",
        "Explanation": "My bubble was mistaken for the whole population.",
        "Reflection Question": "How would you check if most people agree — beyond your circle?",
        "Why Humans Do This": "We're surrounded by similar people, so their views feel like everyone's.",
    },
    "C035": {
        "Statement": "I posted a hot take online and got slammed — I'd assumed 'everyone thinks this.'",
        "Explanation": "My bubble felt like consensus.",
        "Reflection Question": "Online backlash that surprised you because you thought your view was common?",
        "Why Humans Do This": "Our own beliefs feel normal — so we assume others share them.",
    },
    "C036": {
        "Statement": "Two people agreed in a meeting — so the manager assumed the whole team supports the policy.",
        "Explanation": "A vocal few stood in for everyone.",
        "Reflection Question": "When few speak up — what can you really conclude about the rest?",
        "Why Humans Do This": "Loud agreement is easy to notice. Silent disagreement is invisible.",
    },
    "C037": {
        "Statement": "Same mistake from my department gets a pass — from another department, it's a serious problem.",
        "Explanation": "Group membership changed how the mistake was judged.",
        "Reflection Question": "Judge mistakes from 'your people' more gently than from others?",
        "Why Humans Do This": "We extend more trust and leniency to people we see as 'us.'",
    },
    "C038": {
        "Statement": "Our player's foul was accidental. Their player's identical foul was dirty play.",
        "Explanation": "Same action, different story — depending on the team.",
        "Reflection Question": "Would you see it the same way if the teams were swapped?",
        "Why Humans Do This": "Loyalty shapes how we read ambiguous actions — we favor our side.",
    },
    "C039": {
        "Statement": "I excuse a scandal from my party — but I'm outraged by the same thing from the other side.",
        "Explanation": "Group allegiance changed the moral judgment.",
        "Reflection Question": "Would this bother you as much from your own side?",
        "Why Humans Do This": "Protecting our group's reputation can feel like protecting part of ourselves.",
    },
    "C040": {
        "Statement": "Marketing sees lots of personalities in marketing — but assumes everyone in engineering is the same.",
        "Explanation": "Diversity inside my group; sameness assumed outside it.",
        "Reflection Question": "Which groups do you see as diverse — and which do you lump together?",
        "Why Humans Do This": "We know individuals in our group. Other groups we only see from a distance.",
    },
    "C041": {
        "Statement": "I've never lived there — but I'm sure people from that country all think the same on this issue.",
        "Explanation": "Limited exposure became a uniform stereotype.",
        "Reflection Question": "How much real exposure do you have to the group you're generalizing?",
        "Why Humans Do This": "Less detail about a group means broader, simpler assumptions.",
    },
    "C042": {
        "Statement": "A headline says 'young people today' all think one way — but we'd never describe our own generation that simply.",
        "Explanation": "A diverse group was flattened into one attitude.",
        "Reflection Question": "Would you accept this generalization about a group you're in?",
        "Why Humans Do This": "Big categories are easier to write and think about than real diversity.",
    },
    "C043": {
        "Statement": "'90% survival rate' sounds great. '10% mortality rate' sounds scary. Same treatment.",
        "Explanation": "Identical stats, different emotional frame.",
        "Reflection Question": "A decision where wording changed how you felt?",
        "Why Humans Do This": "We react to gain vs loss framing before we sit with the numbers.",
    },
    "C044": {
        "Statement": "'90% fat free' sells better than '10% fat' — same product on the label.",
        "Explanation": "Positive wording beat equivalent negative wording.",
        "Reflection Question": "Chosen a product because of the label, not the content?",
        "Why Humans Do This": "Positive-sounding words trigger a warmer gut reaction — facts aside.",
    },
    "C045": {
        "Statement": "One side says a policy saves 200 lives. The other says it fails to save 100 — same outcome, different frame.",
        "Explanation": "Same numbers, opposite emotional pull.",
        "Reflection Question": "Do you flip the frame to see if the numbers still hold?",
        "Why Humans Do This": "Communicators pick the frame that triggers the feeling they want.",
    },
    "C046": {
        "Statement": "Losing $1,000 hurts more than gaining $1,000 feels good — same amount.",
        "Explanation": "Losses weigh heavier than equal gains.",
        "Reflection Question": "Would you take a 50/50 bet to win or lose $100?",
        "Why Humans Do This": "Avoiding loss mattered more for survival than chasing equal rewards.",
    },
    "C047": {
        "Statement": "In a negotiation, I'll fight harder to keep what I have than to gain something new of equal value.",
        "Explanation": "Losing what we have feels worse than gaining the same amount.",
        "Reflection Question": "Last negotiation — fighting harder to keep or to gain?",
        "Why Humans Do This": "Losses are felt more intensely than equivalent gains — often about twice as much.",
    },
    "C048": {
        "Statement": "I keep a gym membership I never use — canceling feels like losing something.",
        "Explanation": "Avoiding the feeling of loss beats the logic of stopping payment.",
        "Reflection Question": "Keeping something unused — value, or just fear of losing it?",
        "Why Humans Do This": "'Losing' something — even unused — can outweigh the clear case for letting go.",
    },
    "C049": {
        "Statement": "I smoke but tell myself I'm less likely to get sick than other smokers — no real reason why.",
        "Explanation": "'It won't happen to me' without evidence.",
        "Reflection Question": "A risk you've downplayed for yourself but not for others?",
        "Why Humans Do This": "Believing we're the exception reduces anxiety and keeps us moving.",
    },
    "C050": {
        "Statement": "We always estimate projects will finish early — despite a long history of running late.",
        "Explanation": "Past delays get ignored for a hopeful new estimate.",
        "Reflection Question": "Does planning use real history — or just this project's best case?",
        "Why Humans Do This": "We picture the best outcome for our own plans and discount past slip-ups.",
    },
    "C051": {
        "Statement": "I'm not saving for retirement — I'll 'figure it out later' and assume I'll earn more than stats suggest.",
        "Explanation": "A rosier-than-average future assumed without proof.",
        "Reflection Question": "Is your plan based on realistic odds — or a hoped-for best case?",
        "Why Humans Do This": "An optimistic future feels easier than facing uncertain odds.",
    },
    "C052": {
        "Statement": "I stayed in my company's default retirement plan — even though a better one needed just one form.",
        "Explanation": "No action beat a slightly better option.",
        "Reflection Question": "A default choice you've never actually reconsidered?",
        "Why Humans Do This": "Sticking with now takes no effort. Change takes a decision.",
    },
    "C053": {
        "Statement": "We keep outdated software because switching feels risky — though the new system would clearly save time and money.",
        "Explanation": "Familiar beat a rational case for change.",
        "Reflection Question": "Something at work you keep because 'that's how it's always been'?",
        "Why Humans Do This": "Change means uncertainty. Staying put feels safe and predictable.",
    },
    "C054": {
        "Statement": "I've kept my phone's default settings for years — assuming the manufacturer picked the best ones.",
        "Explanation": "Default was treated as best without checking.",
        "Reflection Question": "Default settings you've never questioned — phone, habits, anything?",
        "Why Humans Do This": "Defaults feel like a recommendation. Changing them takes mental effort.",
    },
    "C055": {
        "Statement": "After one online course, I felt like an expert in a field professionals study for years.",
        "Explanation": "A little learning created a lot of confidence.",
        "Reflection Question": "A skill where early confidence later met how much you didn't know?",
        "Why Humans Do This": "Early learning skips the part where you discover how much is still ahead.",
    },
    "C056": {
        "Statement": "As a beginner, I'm sure I could beat a chess grandmaster with the right strategy.",
        "Explanation": "Limited experience hid how wide the skill gap really is.",
        "Reflection Question": "How do you find out how good you really are at something new?",
        "Why Humans Do This": "Without expert knowledge, it's hard to see what expert skill actually looks like.",
    },
    "C057": {
        "Statement": "A few articles online — and I felt more sure about my symptoms than the doctor with thousands of cases.",
        "Explanation": "Surface reading felt like real expertise.",
        "Reflection Question": "When is your own research enough — and when do you need a pro?",
        "Why Humans Do This": "A little information can feel complete when depth is still missing.",
    },
    "C058": {
        "Statement": "I trusted a diet tip from my doctor — even though it wasn't their area of expertise.",
        "Explanation": "General authority stretched to a specific claim.",
        "Reflection Question": "Does their authority actually match this specific advice?",
        "Why Humans Do This": "Titles and status become shortcuts for trust — instead of checking each claim.",
    },
    "C059": {
        "Statement": "I didn't question a flawed plan because a senior executive proposed it.",
        "Explanation": "Rank stood in for a sound argument.",
        "Reflection Question": "Gone along with an idea mainly because of who suggested it?",
        "Why Humans Do This": "Deferring to authority feels safer than scrutinizing the idea itself.",
    },
    "C060": {
        "Statement": "An actor who played a doctor on TV comments on real medicine — and viewers partly trust them because of the role.",
        "Explanation": "A fictional role lent unearned credibility.",
        "Reflection Question": "Would you trust this take as much if you didn't know who was speaking?",
        "Why Humans Do This": "Familiarity with an authority-like role can feel like real expertise.",
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb["Challenge Library"]
    headers = [cell.value for cell in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    id_col = idx["challenge_id"]
    updated = 0

    for row in ws.iter_rows(min_row=2):
        cid = row[id_col - 1].value
        if cid not in UPDATES:
            continue
        for field, value in UPDATES[cid].items():
            row[idx[field] - 1].value = value
        updated += 1

    wb.save(SOURCE)
    print(f"Updated {updated} challenges in Excel")


if __name__ == "__main__":
    main()
